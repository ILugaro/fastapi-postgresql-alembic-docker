"""Парсинг локальных файлов"""
import re

from openpyxl import load_workbook

from app import constants
from app.models import Category, CategoryPath, City, Nomenclature
from app.models.product import Product


class ParserService:
    """Обработка Excel документа"""

    @staticmethod
    async def create_products_from_excel(path_of_file: str) -> list[Product]:
        """
        Получение данных о продукции из xlsx файла

        :param path_of_file: Путь к xlsx файлу с продуктами
        :return: Объекты продуктов из указанного excel файла
        """
        wb = load_workbook(path_of_file)
        sheet = wb.active

        products: list[Product] = []
        rows = []
        cities = set()
        categores = set()
        category_paths = set()
        nomenclatures = set()

        for row in sheet.iter_rows(min_row=constants.FIRST_DATA_ROW_IN_XLSX, values_only=True):
            cities.add(row[constants.CITY_COL - 1])
            categores.add(row[constants.CATEGORY_COL - 1])
            category_paths.add(row[constants.CATEGORY_PATH_COL - 1])
            nomenclatures.add(row[constants.NOMENCLATURE_COL - 1])

            rows.append(row)

        category_bases = {category: Category(category=category) for category in categores}
        city_bases = {city: City(city=city) for city in cities}
        category_path_bases = {
            category_path: CategoryPath(category_path=category_path) for category_path in category_paths
        }
        nomenclature_bases = {
            nomenclature: Nomenclature(nomenclature=nomenclature) for nomenclature in nomenclatures
        }

        for row in rows:

            # проверки наличия и формата URL
            if not row[constants.URL_COL - 1]:
                raise Exception(constants.NOT_URL.format(row[constants.NAME_COL - 1]))
            if not re.fullmatch(constants.URL_PATH_FOR_PRODUCT + r'/\d*/', row[constants.URL_COL - 1]):
                raise Exception(
                    constants.NOT_EXPECTED_URL.format(
                        row[constants.URL_COL - 1], row[constants.NAME_COL - 1]
                    )
                )

            products.append(
                Product(
                    name=row[constants.NAME_COL - 1],
                    articul=row[constants.ARTICUL_COL - 1],
                    gost=row[constants.GOST_COL - 1],
                    brand=row[constants.BRAND_COL - 1],
                    category=category_bases[row[constants.CATEGORY_COL - 1]],
                    nomenclature=nomenclature_bases[row[constants.NOMENCLATURE_COL - 1]],
                    category_path=category_path_bases[row[constants.CATEGORY_PATH_COL - 1]],
                    price=row[constants.PRICE_COL - 1],
                    url_id=row[constants.URL_COL - 1].split('/')[-2],
                    warehouse=row[constants.WAREHOUSE_COL - 1],
                    instock=row[constants.INSTOCK_COL - 1],
                    count=row[constants.COUNT - 1],
                    city=city_bases[row[constants.CITY_COL - 1]],
                    updated_at=row[constants.UPDATED_AT_COL - 1],
                    discount_price=row[constants.DISCOUNT_PRICE_COL - 1],
                    razmer=row[constants.RAZMER_COL - 1],
                )
            )
        return products
